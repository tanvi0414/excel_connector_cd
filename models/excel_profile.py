import base64
from io import BytesIO
from datetime import date, datetime, time, timedelta

from odoo import _, api, fields, models
from odoo.exceptions import UserError, ValidationError
from odoo.tools.safe_eval import safe_eval

try:
    import xlsxwriter
except Exception:  # pragma: no cover
    xlsxwriter = None


class ExcelProfile(models.Model):
    _name = 'nl.excel.profile'
    _description = 'Excel Connector Profile'
    _order = 'sequence, name, id'

    name = fields.Char(required=True)
    active = fields.Boolean(default=True)
    sequence = fields.Integer(default=10)
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company, index=True)
    run_as_user_id = fields.Many2one('res.users', string='Run As', default=lambda self: self.env.user, required=True)
    preset_key = fields.Selection(selection=[
        ('customer_export', 'Customers'),
        ('product_export', 'Products'),
        ('sale_order_export', 'Sales Orders'),
        ('invoice_export', 'Customer Invoices'),
        ('purchase_order_export', 'Purchase Orders'),
        ('stock_quant_export', 'Inventory'),
    ], string='Preset')
    model_id = fields.Many2one('ir.model', string='Model', required=True, ondelete='cascade')
    model_name = fields.Char(related='model_id.model', store=True)
    mode = fields.Selection([
        ('both', 'Import + Export'),
        ('export', 'Export Only'),
        ('import', 'Import Only'),
    ], default='both', required=True)
    sheet_name = fields.Char(default='Sheet1', required=True)
    export_filename = fields.Char(default='odoo_export.xlsx', required=True)
    include_headers = fields.Boolean(default=True)
    domain = fields.Text(default='[]')
    order_by = fields.Char(default='id desc')
    limit = fields.Integer(default=0)
    import_mode = fields.Selection([
        ('create', 'Create New'),
        ('update', 'Update Existing'),
        ('upsert', 'Update or Create'),
    ], default='upsert', required=True)
    key_field_id = fields.Many2one(
        'ir.model.fields',
        string='Match Using',
        domain="[('model_id', '=', model_id), ('store', '=', True), ('ttype', 'not in', ['one2many', 'many2many'])]",
        ondelete='set null',
        help='Used for updates and upserts. Example: Email, Internal Reference, or External ID field.',
    )
    stop_on_error = fields.Boolean(default=False)
    auto_export = fields.Boolean(string='Scheduled Export', default=False)
    interval_number = fields.Integer(default=1)
    interval_type = fields.Selection([
        ('hours', 'Hours'),
        ('days', 'Days'),
        ('weeks', 'Weeks'),
    ], default='days', required=True)
    next_run_at = fields.Datetime(string='Next Scheduled Run')
    last_run_at = fields.Datetime(readonly=True)
    last_attachment_id = fields.Many2one('ir.attachment', string='Last Export File', ondelete='set null', readonly=True)
    line_ids = fields.One2many('nl.excel.profile.field', 'profile_id', string='Columns', copy=True)
    log_ids = fields.One2many('nl.excel.log', 'profile_id', string='Logs')
    log_count = fields.Integer(compute='_compute_log_count')

    @api.depends('log_ids')
    def _compute_log_count(self):
        for profile in self:
            profile.log_count = len(profile.log_ids)

    @api.constrains('interval_number')
    def _check_interval_number(self):
        for profile in self:
            if profile.interval_number < 1:
                raise ValidationError(_('Schedule interval must be at least 1.'))

    @api.constrains('line_ids')
    def _check_line_ids(self):
        for profile in self:
            if not profile.line_ids:
                continue
            field_names = profile.line_ids.mapped('field_id.name')
            if len(field_names) != len(set(field_names)):
                raise ValidationError(_('Each field can only be added once in a profile.'))

    def name_get(self):
        result = []
        for rec in self:
            model_name = rec.model_id.name or rec.model_name or ''
            result.append((rec.id, f"{rec.name} · {model_name}" if model_name else rec.name))
        return result

    def action_open_logs(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Sync Logs'),
            'res_model': 'nl.excel.log',
            'view_mode': 'list,form',
            'domain': [('profile_id', '=', self.id)],
            'context': {'default_profile_id': self.id},
        }

    def action_apply_preset(self):
        self.ensure_one()
        if not self.preset_key:
            raise UserError(_('Select a preset first.'))
        preset = self._get_presets().get(self.preset_key)
        if not preset:
            raise UserError(_('Preset not found.'))

        model = self.env['ir.model'].search([('model', '=', preset['model'])], limit=1)
        if not model:
            raise UserError(_('Model %s was not found in this database.') % preset['model'])

        self.write({
            'model_id': model.id,
            'sheet_name': preset.get('sheet_name', 'Sheet1'),
            'export_filename': preset.get('filename', 'odoo_export.xlsx'),
            'domain': preset.get('domain', '[]'),
            'order_by': preset.get('order_by', 'id desc'),
        })
        self.line_ids.unlink()
        line_commands = []
        for sequence, field_name in enumerate(preset['fields'], start=1):
            field_rec = self.env['ir.model.fields'].search([
                ('model', '=', preset['model']),
                ('name', '=', field_name),
                ('store', '=', True),
            ], limit=1)
            if not field_rec:
                continue
            line_commands.append((0, 0, {
                'sequence': sequence * 10,
                'field_id': field_rec.id,
                'column_label': field_rec.field_description or field_name,
                'required_on_import': False,
            }))
        self.write({'line_ids': line_commands})
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Preset Applied'),
                'message': _('The profile has been filled with business-ready columns.'),
                'type': 'success',
                'sticky': False,
            }
        }

    def action_download_template(self):
        self.ensure_one()
        content = self._build_template_workbook()
        attachment = self._create_attachment(
            name=self._template_filename(),
            content=content,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self._create_log(
            direction='export',
            state='success',
            rows_processed=0,
            rows_failed=0,
            message=_('Blank template generated for import/export use.'),
            attachment=attachment,
        )
        return self._download_attachment_action(attachment)

    def action_run_export(self):
        self.ensure_one()
        attachment = self._run_export_create_attachment()
        return self._download_attachment_action(attachment)

    def action_open_import_wizard(self):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_window',
            'name': _('Import Excel File'),
            'res_model': 'nl.excel.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {'default_profile_id': self.id},
        }

    def cron_run_scheduled_exports(self):
        profiles = self.search([
            ('active', '=', True),
            ('auto_export', '=', True),
            ('next_run_at', '!=', False),
            ('next_run_at', '<=', fields.Datetime.now()),
        ])
        for profile in profiles:
            try:
                profile.with_user(profile.run_as_user_id).with_company(profile.company_id)._run_export_create_attachment(scheduled=True)
            except Exception as exc:  # pragma: no cover - defensive logging for cron
                profile._create_log(
                    direction='export',
                    state='error',
                    rows_processed=0,
                    rows_failed=0,
                    message=str(exc),
                )
                profile._update_next_run()

    def _run_export_create_attachment(self, scheduled=False):
        self.ensure_one()
        content, row_count = self._build_export_workbook()
        attachment = self._create_attachment(
            name=self._resolved_export_filename(),
            content=content,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.write({
            'last_run_at': fields.Datetime.now(),
            'last_attachment_id': attachment.id,
        })
        self._update_next_run()
        self._create_log(
            direction='export',
            state='success',
            rows_processed=row_count,
            rows_failed=0,
            message=_('Excel export completed successfully.') if not scheduled else _('Scheduled Excel export completed successfully.'),
            attachment=attachment,
        )
        return attachment

    def _update_next_run(self):
        self.ensure_one()
        if not self.auto_export:
            return
        base_dt = fields.Datetime.now()
        delta = timedelta(hours=self.interval_number)
        if self.interval_type == 'days':
            delta = timedelta(days=self.interval_number)
        elif self.interval_type == 'weeks':
            delta = timedelta(weeks=self.interval_number)
        self.next_run_at = base_dt + delta

    def _create_attachment(self, name, content, mimetype):
        self.ensure_one()
        return self.env['ir.attachment'].create({
            'name': name,
            'datas': base64.b64encode(content),
            'mimetype': mimetype,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })

    def _download_attachment_action(self, attachment):
        self.ensure_one()
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def _resolved_export_filename(self):
        self.ensure_one()
        name = self.export_filename or f'{self.name}.xlsx'
        if not name.lower().endswith('.xlsx'):
            name += '.xlsx'
        return name

    def _template_filename(self):
        self.ensure_one()
        base = (self.name or 'template').strip().replace('/', '_')
        return f'{base}_template.xlsx'

    def _safe_domain(self):
        self.ensure_one()
        domain_text = self.domain or '[]'
        try:
            value = safe_eval(domain_text, {'uid': self.env.uid})
        except Exception as exc:
            raise UserError(_('Invalid domain: %s') % exc)
        if not isinstance(value, (list, tuple)):
            raise UserError(_('Domain must evaluate to a list or tuple.'))
        return list(value)

    def _build_export_workbook(self):
        self.ensure_one()
        if xlsxwriter is None:
            raise UserError(_('xlsxwriter is not available on this server.'))
        if not self.line_ids:
            raise UserError(_('Add at least one column before exporting.'))

        records = self.env[self.model_name].search(
            self._safe_domain(),
            order=self.order_by or 'id desc',
            limit=self.limit or None,
        )

        buffer = BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        sheet_name = (self.sheet_name or 'Sheet1')[:31]
        worksheet = workbook.add_worksheet(sheet_name)
        header_format = workbook.add_format({'bold': True, 'bg_color': '#D9EAF7', 'border': 1})
        normal_format = workbook.add_format({'text_wrap': False})

        lines = self.line_ids.sorted('sequence')
        if self.include_headers:
            for col, line in enumerate(lines):
                worksheet.write(0, col, line.column_label or line.field_id.field_description or line.field_id.name, header_format)
            worksheet.freeze_panes(1, 0)
            worksheet.autofilter(0, 0, 0, max(len(lines) - 1, 0))
            start_row = 1
        else:
            start_row = 0

        for row_index, record in enumerate(records, start=start_row):
            for col, line in enumerate(lines):
                worksheet.write(row_index, col, self._export_cell_value(record, line), normal_format)

        for col, line in enumerate(lines):
            width = max(14, min(40, len(line.column_label or line.field_id.field_description or line.field_id.name) + 4))
            worksheet.set_column(col, col, width)

        workbook.close()
        buffer.seek(0)
        return buffer.read(), len(records)

    def _build_template_workbook(self):
        self.ensure_one()
        if xlsxwriter is None:
            raise UserError(_('xlsxwriter is not available on this server.'))
        if not self.line_ids:
            raise UserError(_('Add at least one column before generating a template.'))

        buffer = BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet((self.sheet_name or 'Sheet1')[:31])
        header_format = workbook.add_format({'bold': True, 'bg_color': '#E8F5E9', 'border': 1})
        note_format = workbook.add_format({'italic': True, 'font_color': '#666666'})

        lines = self.line_ids.sorted('sequence')
        for col, line in enumerate(lines):
            label = line.column_label or line.field_id.field_description or line.field_id.name
            worksheet.write(0, col, label, header_format)
            hint = self._field_template_hint(line.field_id)
            worksheet.write(1, col, hint, note_format)
            worksheet.set_column(col, col, max(16, min(42, len(label) + 6)))

        worksheet.freeze_panes(2, 0)
        workbook.close()
        buffer.seek(0)
        return buffer.read()

    def _field_template_hint(self, field_rec):
        hints = {
            'char': 'Text',
            'text': 'Long text',
            'html': 'Text',
            'integer': 'Whole number',
            'float': 'Decimal number',
            'monetary': 'Decimal number',
            'boolean': 'TRUE / FALSE',
            'date': 'Date',
            'datetime': 'Date and time',
            'many2one': 'Record name or ID',
            'selection': 'Internal value or label',
        }
        return hints.get(field_rec.ttype, field_rec.ttype or '')

    def _export_cell_value(self, record, line):
        field_name = line.field_id.name
        field = record._fields[field_name]
        value = record[field_name]
        if not value:
            return ''
        if field.type == 'many2one':
            return value.display_name
        if field.type in ('one2many', 'many2many'):
            return ', '.join(value.mapped('display_name'))
        if field.type == 'boolean':
            return 'TRUE' if value else 'FALSE'
        if field.type == 'date':
            return fields.Date.to_string(value)
        if field.type == 'datetime':
            return fields.Datetime.to_string(value)
        if field.type == 'selection':
            selection = dict(field.selection(record.env)) if callable(field.selection) else dict(field.selection or [])
            return selection.get(value, value)
        return value

    def _parse_import_value(self, field_rec, cell_value):
        if cell_value in (None, ''):
            return False
        ttype = field_rec.ttype
        if ttype in ('char', 'text', 'html'):
            return str(cell_value).strip()
        if ttype == 'integer':
            return int(cell_value)
        if ttype in ('float', 'monetary'):
            return float(cell_value)
        if ttype == 'boolean':
            if isinstance(cell_value, bool):
                return cell_value
            text = str(cell_value).strip().lower()
            return text in ('1', 'true', 'yes', 'y')
        if ttype == 'date':
            if isinstance(cell_value, datetime):
                return cell_value.date()
            if isinstance(cell_value, date):
                return cell_value
            return fields.Date.to_date(str(cell_value))
        if ttype == 'datetime':
            if isinstance(cell_value, datetime):
                return cell_value
            if isinstance(cell_value, date):
                return datetime.combine(cell_value, time.min)
            return fields.Datetime.to_datetime(str(cell_value))
        if ttype == 'many2one':
            relation_model = self.env[field_rec.relation]
            if isinstance(cell_value, (int, float)) and float(cell_value).is_integer():
                rec = relation_model.browse(int(cell_value)).exists()
                if rec:
                    return rec.id
            rec = relation_model.search([(relation_model._rec_name or 'name', '=', str(cell_value).strip())], limit=1)
            if rec:
                return rec.id
            raise UserError(_('Could not match "%s" for field "%s".') % (cell_value, field_rec.field_description or field_rec.name))
        if ttype == 'selection':
            value = str(cell_value).strip()
            field = self.env[self.model_name]._fields[field_rec.name]
            options = field.selection(self.env) if callable(field.selection) else field.selection
            option_map = {k: k for k, _label in options or []}
            option_map.update({label: key for key, label in options or []})
            if value in option_map:
                return option_map[value]
            raise UserError(_('Invalid selection value "%s" for field "%s".') % (value, field_rec.field_description or field_rec.name))
        return cell_value

    def _import_rows(self, workbook_bytes, dry_run=False, filename=None):
        self.ensure_one()
        try:
            from openpyxl import load_workbook
        except Exception as exc:  # pragma: no cover
            raise UserError(_('Excel import needs the Python package openpyxl on the server. Original error: %s') % exc)

        if not self.line_ids:
            raise UserError(_('Add at least one mapped column before importing.'))

        workbook = load_workbook(BytesIO(workbook_bytes), data_only=True)
        if self.sheet_name and self.sheet_name in workbook.sheetnames:
            worksheet = workbook[self.sheet_name]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            raise UserError(_('The uploaded workbook is empty.'))

        headers = [str(cell).strip() if cell not in (None, '') else '' for cell in rows[0]]
        if not any(headers):
            raise UserError(_('The first row must contain column headers.'))

        mapped_lines = {}
        for line in self.line_ids:
            keys = {
                (line.column_label or '').strip().lower(),
                (line.field_id.field_description or '').strip().lower(),
                line.field_id.name.strip().lower(),
            }
            for key in keys:
                if key:
                    mapped_lines[key] = line

        column_map = {}
        for index, header in enumerate(headers):
            line = mapped_lines.get(header.lower())
            if line:
                column_map[index] = line

        if not column_map:
            raise UserError(_('None of the uploaded columns match the profile mapping.'))

        model = self.env[self.model_name]
        processed = 0
        failed = 0
        errors = []

        for row_number, row in enumerate(rows[1:], start=2):
            if not any(value not in (None, '') for value in row):
                continue
            try:
                vals = {}
                for index, line in column_map.items():
                    cell_value = row[index] if index < len(row) else None
                    parsed = self._parse_import_value(line.field_id, cell_value)
                    if parsed is False and line.required_on_import:
                        raise UserError(_('Column "%s" is required.') % (line.column_label or line.field_id.field_description or line.field_id.name))
                    if parsed is not False:
                        vals[line.field_id.name] = parsed
                if not vals:
                    continue

                record = self._find_matching_record_for_import(model, vals)
                if dry_run:
                    if self.import_mode == 'update' and not record:
                        raise UserError(_('No existing record found for update.'))
                else:
                    if self.import_mode == 'create':
                        model.create(vals)
                    elif self.import_mode == 'update':
                        if not record:
                            raise UserError(_('No existing record found for update.'))
                        record.write(vals)
                    else:  # upsert
                        if record:
                            record.write(vals)
                        else:
                            model.create(vals)
                processed += 1
            except Exception as exc:
                failed += 1
                errors.append({'row_number': row_number, 'row_values': list(row), 'error': str(exc)})
                if self.stop_on_error:
                    break

        error_attachment = None
        if errors:
            error_attachment = self._build_error_attachment(headers, errors, filename or 'import_errors.xlsx')

        state = 'success'
        if failed and processed:
            state = 'warning'
        elif failed and not processed:
            state = 'error'

        message = _('Processed %(processed)s row(s), failed %(failed)s row(s).', processed=processed, failed=failed)
        if dry_run:
            message = _('Dry run complete. ') + message

        self._create_log(
            direction='import',
            state=state,
            rows_processed=processed,
            rows_failed=failed,
            message=message,
            attachment=error_attachment,
        )

        return {
            'processed': processed,
            'failed': failed,
            'message': message,
            'attachment_id': error_attachment.id if error_attachment else False,
        }

    def _find_matching_record_for_import(self, model, vals):
        self.ensure_one()
        if self.import_mode == 'create' or not self.key_field_id:
            return model.browse()
        key_name = self.key_field_id.name
        if key_name not in vals:
            raise UserError(_('The match field "%s" is missing in the uploaded file.') % (self.key_field_id.field_description or key_name))
        return model.search([(key_name, '=', vals[key_name])], limit=1)

    def _build_error_attachment(self, headers, errors, filename):
        self.ensure_one()
        if xlsxwriter is None:
            return False
        buffer = BytesIO()
        workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
        worksheet = workbook.add_worksheet('Errors')
        header_fmt = workbook.add_format({'bold': True, 'bg_color': '#FDECEC', 'border': 1})
        for index, header in enumerate(headers + ['Error']):
            worksheet.write(0, index, header, header_fmt)
        for row_index, error in enumerate(errors, start=1):
            row_values = error['row_values']
            for col_index, value in enumerate(row_values):
                worksheet.write(row_index, col_index, '' if value is None else value)
            worksheet.write(row_index, len(headers), error['error'])
        workbook.close()
        buffer.seek(0)
        safe_name = (filename or 'import_errors.xlsx').replace('.xlsx', '_errors.xlsx')
        return self._create_attachment(
            name=safe_name,
            content=buffer.read(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

    def _create_log(self, direction, state, rows_processed, rows_failed, message, attachment=False):
        self.ensure_one()
        self.env['nl.excel.log'].create({
            'name': f'{self.name} - {direction.title()}',
            'profile_id': self.id,
            'company_id': self.company_id.id,
            'direction': direction,
            'state': state,
            'rows_processed': rows_processed,
            'rows_failed': rows_failed,
            'message': message,
            'attachment_id': attachment.id if attachment else False,
        })

    def _get_presets(self):
        return {
            'customer_export': {
                'model': 'res.partner',
                'sheet_name': 'Customers',
                'filename': 'customers.xlsx',
                'order_by': 'id desc',
                'fields': ['name', 'email', 'phone', 'city', 'country_id', 'vat'],
            },
            'product_export': {
                'model': 'product.product',
                'sheet_name': 'Products',
                'filename': 'products.xlsx',
                'order_by': 'id desc',
                'fields': ['default_code', 'name', 'barcode', 'lst_price', 'standard_price', 'categ_id', 'active'],
            },
            'sale_order_export': {
                'model': 'sale.order',
                'sheet_name': 'Sales Orders',
                'filename': 'sales_orders.xlsx',
                'order_by': 'id desc',
                'fields': ['name', 'date_order', 'partner_id', 'user_id', 'state', 'amount_total', 'currency_id'],
            },
            'invoice_export': {
                'model': 'account.move',
                'sheet_name': 'Invoices',
                'filename': 'customer_invoices.xlsx',
                'domain': "[('move_type', '=', 'out_invoice')]",
                'order_by': 'invoice_date desc, id desc',
                'fields': ['name', 'invoice_date', 'partner_id', 'state', 'payment_state', 'amount_total', 'currency_id'],
            },
            'purchase_order_export': {
                'model': 'purchase.order',
                'sheet_name': 'Purchase Orders',
                'filename': 'purchase_orders.xlsx',
                'order_by': 'id desc',
                'fields': ['name', 'date_order', 'partner_id', 'user_id', 'state', 'amount_total', 'currency_id'],
            },
            'stock_quant_export': {
                'model': 'stock.quant',
                'sheet_name': 'Inventory',
                'filename': 'inventory.xlsx',
                'order_by': 'id desc',
                'fields': ['product_id', 'location_id', 'quantity', 'available_quantity', 'inventory_quantity_set'],
            },
        }


class ExcelProfileField(models.Model):
    _name = 'nl.excel.profile.field'
    _description = 'Excel Connector Profile Field'
    _order = 'sequence, id'

    sequence = fields.Integer(default=10)
    profile_id = fields.Many2one('nl.excel.profile', string='Profile', required=True, ondelete='cascade')
    model_id = fields.Many2one(related='profile_id.model_id', store=True)
    field_id = fields.Many2one(
        'ir.model.fields',
        string='Field',
        required=True,
        domain="[('model_id', '=', model_id), ('store', '=', True)]",
        ondelete='cascade',
    )
    column_label = fields.Char(required=True)
    required_on_import = fields.Boolean(string='Required on Import', default=False)

    @api.onchange('field_id')
    def _onchange_field_id(self):
        for rec in self:
            if rec.field_id and not rec.column_label:
                rec.column_label = rec.field_id.field_description or rec.field_id.name
